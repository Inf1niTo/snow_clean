import os
import sys
import pandas as pd
from openpyxl import load_workbook
from copy import copy


# =============================
# Определяем папку запуска
# =============================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

input_file = os.path.join(BASE_DIR, "выгрузка.xlsx")
template_file = os.path.join(BASE_DIR, "Tamplate.xlsx")
output_folder = os.path.join(BASE_DIR, "Отчеты_по_округам")

os.makedirs(output_folder, exist_ok=True)

# =============================
# Проверка наличия файлов
# =============================
if not os.path.exists(input_file):
    print("❌ Файл 'выгрузка.xlsx' не найден рядом с программой")
    input("Нажмите Enter для выхода...")
    sys.exit()

if not os.path.exists(template_file):
    print("❌ Файл 'Tamplate.xlsx' не найден рядом с программой")
    input("Нажмите Enter для выхода...")
    sys.exit()


# =============================
# Чтение данных
# =============================
df = pd.read_excel(input_file)

numeric_cols = [
    "Не подлежит контролю",
    "Иные статусы",
    "В статусе «Исправно»",
    "Техника вышла на уборку по данным СОК"
]

df[numeric_cols] = df[numeric_cols].fillna(0)


# =============================
# Агрегация
# =============================
grouped = (
    df.groupby(["Округ", "Балансодержатель"])
      .agg(
          Всего=("№ п/п", "count"),
          Не_подлежит=("Не подлежит контролю", "sum"),
          Иные=("Иные статусы", "sum"),
          Исправно=("В статусе «Исправно»", "sum"),
          Вышла=("Техника вышла на уборку по данным СОК", "sum")
      )
      .reset_index()
)


# =============================
# Формирование отчетов
# =============================
for okrug in df["Округ"].unique():

    wb = load_workbook(template_file)

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws = wb.active
    ws.title = "Свод"

    df_okrug = grouped[grouped["Округ"] == okrug]
    df_raw = df[df["Округ"] == okrug]

    start_row = 5
    template_row_idx = start_row

    ws["B1"] = f"Статистика по выходу техники по данным СОК по {okrug}"

    # ===== Поиск строки ВСЕГО =====
    total_row_template = None
    for row_idx in range(start_row, ws.max_row + 1):
        if ws[f"B{row_idx}"].value == "ВСЕГО":
            total_row_template = row_idx
            break

    # ===== Сохранение стиля шаблонной строки =====
    template_style = {}
    for col in range(2, 9):
        cell = ws.cell(row=template_row_idx, column=col)
        template_style[col] = {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
        }

    template_merges = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == template_row_idx:
            template_merges.append((merged.min_col, merged.max_col))

    # ===== Удаление старых строк =====
    if total_row_template:
        ws.delete_rows(start_row, total_row_template - start_row)

    # ===== Вставка данных =====
    for i, (_, row_data) in enumerate(df_okrug.iterrows()):
        r = start_row + i
        ws.insert_rows(r)

        # Применяем стиль
        for col in range(2, 9):
            new_cell = ws.cell(row=r, column=col)
            style = template_style[col]

            new_cell.font = copy(style["font"])
            new_cell.fill = copy(style["fill"])
            new_cell.border = copy(style["border"])
            new_cell.alignment = copy(style["alignment"])
            new_cell.number_format = style["number_format"]
            new_cell.protection = copy(style["protection"])

        # Заполняем данные
        ws[f"B{r}"] = row_data["Балансодержатель"]
        ws[f"C{r}"] = row_data["Всего"]
        ws[f"D{r}"] = row_data["Не_подлежит"]
        ws[f"E{r}"] = row_data["Иные"]
        ws[f"F{r}"] = row_data["Исправно"]
        ws[f"G{r}"] = row_data["Вышла"]

        ws[f"H{r}"] = f"=G{r}/F{r}"
        ws[f"H{r}"].number_format = "0%"

        # Восстановление объединений
        for min_col, max_col in template_merges:
            ws.merge_cells(
                start_row=r,
                start_column=min_col,
                end_row=r,
                end_column=max_col
            )

    last_data_row = start_row + len(df_okrug) - 1
    total_row = last_data_row + 1

    # ===== Копирование стиля строки ВСЕГО =====
    if total_row_template:
        for col in range(2, 9):
            template_cell = ws.cell(row=total_row_template, column=col)
            new_cell = ws.cell(row=total_row, column=col)

            new_cell.font = copy(template_cell.font)
            new_cell.fill = copy(template_cell.fill)
            new_cell.border = copy(template_cell.border)
            new_cell.alignment = copy(template_cell.alignment)
            new_cell.number_format = template_cell.number_format
            new_cell.protection = copy(template_cell.protection)

    # ===== Формулы строки ВСЕГО =====
    ws[f"B{total_row}"] = "ВСЕГО"
    ws[f"C{total_row}"] = f"=SUM(C{start_row}:C{last_data_row})"
    ws[f"D{total_row}"] = f"=SUM(D{start_row}:D{last_data_row})"
    ws[f"E{total_row}"] = f"=SUM(E{start_row}:E{last_data_row})"
    ws[f"F{total_row}"] = f"=SUM(F{start_row}:F{last_data_row})"
    ws[f"G{total_row}"] = f"=SUM(G{start_row}:G{last_data_row})"
    ws[f"H{total_row}"] = f"=G{total_row}/F{total_row}"
    ws[f"H{total_row}"].number_format = "0%"

    # ===== Лист выгрузки =====
    ws_raw = wb.create_sheet("Выгрузка")

    for col_idx, col_name in enumerate(df_raw.columns, start=1):
        ws_raw.cell(row=1, column=col_idx, value=col_name)

    for r_idx, row_data in enumerate(df_raw.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row_data, start=1):
            ws_raw.cell(row=r_idx, column=c_idx, value=value)

    ws_raw.auto_filter.ref = ws_raw.dimensions

    for column_cells in ws_raw.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws_raw.column_dimensions[column_cells[0].column_letter].width = length + 2

    safe_name = okrug.replace("/", "-")
    wb.save(os.path.join(output_folder, f"{safe_name}.xlsx"))


print("\n✅ Готово! Файлы созданы в папке 'Отчеты_по_округам'")
input("Нажмите Enter для выхода...")
